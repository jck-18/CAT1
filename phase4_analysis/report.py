"""Phase 4b - Reporting.

Takes the findings produced by analyze_security.py (plus the raw Phase 1/2/3
outputs they were derived from) and produces the deliverables: a set of PNG
charts for the slides and a multi-sheet Excel workbook that is the assessment
report itself.

Inputs:
  outputs/findings.json                       (run analyze_security.py first)
  phase1_discovery/outputs/hosts.json
  phase2_capture/outputs/protocol_stats.json
  phase3_spoofing/outputs/mac_log.json        (optional)

Produces:
  outputs/charts/*.png   - findings by severity, open ports per host, top
                           services, protocol distribution, top talkers
  outputs/report.xlsx    - Summary / Findings / Hosts / Open Ports /
                           Protocol Stats / Top Talkers / Evidence / MAC Log /
                           Firewall Rules / Charts

Usage:
    python report.py
    python report.py --no-charts     # workbook only
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shared import config
from shared.utils import (
    banner, die, ensure_dir, human_bytes, info, ok, read_json, rel, step,
    truncate, warn,
)

try:
    import matplotlib
    matplotlib.use("Agg")            # no GUI: we only ever write files
    import matplotlib.pyplot as plt
    import pandas as pd
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    die(f"a reporting dependency is missing ({exc.name}).\n"
        "    Run:  pip install -r requirements.txt")

# Muted, print-safe palette - these end up on a projector.
SEVERITY_COLOURS = {
    "Critical": "#8E1B1B",
    "High": "#C4622D",
    "Medium": "#D9A404",
    "Low": "#5B8C5A",
    "Info": "#7A8B99",
}
SEVERITY_FILLS = {
    "Critical": "FFC7CE",
    "High": "FFD8A8",
    "Medium": "FFF2CC",
    "Low": "E2EFDA",
    "Info": "F2F2F2",
}
SEVERITY_ORDER = ["Critical", "High", "Medium", "Low", "Info"]
ACCENT = "#2F4858"


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------


def load_everything() -> dict:
    if not config.FINDINGS_JSON.exists():
        die(f"{rel(config.FINDINGS_JSON)} not found.\n"
            "    Run 'python analyze_security.py' first (add --sample if "
            "Phases 1 and 2 have not delivered data yet).")

    findings_doc = read_json(config.FINDINGS_JSON)
    summary = findings_doc.get("summary", {})
    sources = summary.get("sources", {})

    # analyze_security.py records where it read from, so the report follows the
    # same inputs - including the synthetic ones when --sample was used.
    if sources.get("sample"):
        warn("this report is built from SAMPLE data - regenerate it once "
             "Phases 1 and 2 deliver real outputs")
        base = config.PHASE4_OUTPUTS / "sample_inputs"
        hosts_path = base / "hosts.json"
        stats_path = base / "protocol_stats.json"
    else:
        hosts_path = config.HOSTS_JSON
        stats_path = config.PROTOCOL_STATS_JSON

    def optional(path: Path):
        if not path.exists():
            warn(f"{rel(path)} not found - the related sheets will be empty")
            return None
        return read_json(path)

    return {
        "summary": summary,
        "findings": findings_doc.get("findings", []),
        "hosts": optional(hosts_path) or [],
        "stats": optional(stats_path) or {},
        "mac_log": read_json(config.MAC_LOG_JSON)
                   if config.MAC_LOG_JSON.exists() else None,
        "is_sample": bool(sources.get("sample")),
    }


# --------------------------------------------------------------------------
# Frames
# --------------------------------------------------------------------------


def findings_frame(findings: list[dict]) -> "pd.DataFrame":
    frame = pd.DataFrame(findings)
    if frame.empty:
        return pd.DataFrame(columns=["id", "severity", "category", "phase",
                                     "title", "asset", "evidence", "risk",
                                     "recommendation"])
    frame["severity"] = pd.Categorical(frame["severity"],
                                       categories=SEVERITY_ORDER, ordered=True)
    return frame.sort_values(["severity", "asset"]).reset_index(drop=True)


def hosts_frame(hosts: list[dict]) -> "pd.DataFrame":
    rows = []
    for host in hosts:
        open_ports = [p for p in host.get("ports", []) if p.get("state") == "open"]
        rows.append({
            "ip": host.get("ip"),
            "team_host": config.host_label(host.get("ip"), mac=host.get("mac")) or "",
            "hostname": host.get("hostname"),
            "mac": host.get("mac"),
            "vendor": host.get("vendor"),
            "os": host.get("os"),
            "os_confidence": host.get("os_accuracy"),
            "state": host.get("state"),
            "open_ports": len(open_ports),
            "ports": ", ".join(str(p["port"]) for p in open_ports),
        })
    return pd.DataFrame(rows)


def ports_frame(hosts: list[dict]) -> "pd.DataFrame":
    rows = []
    for host in hosts:
        for port in host.get("ports", []):
            rows.append({
                "ip": host.get("ip"),
                "team_host": config.host_label(host.get("ip"), mac=host.get("mac")) or "",
                "port": port.get("port"),
                "protocol": port.get("protocol"),
                "service": port.get("service"),
                "version": port.get("version"),
                "state": port.get("state"),
            })
    frame = pd.DataFrame(rows)
    return frame.sort_values(["ip", "port"]) if not frame.empty else frame


def protocol_frame(stats: dict) -> "pd.DataFrame":
    counts = stats.get("protocol_counts", {})
    byte_counts = stats.get("protocol_bytes", {})
    total = sum(counts.values()) or 1
    rows = [{
        "protocol": protocol,
        "packets": count,
        "share_percent": round(100 * count / total, 2),
        "bytes": byte_counts.get(protocol, 0),
        "bytes_human": human_bytes(byte_counts.get(protocol, 0)),
    } for protocol, count in counts.items()]
    return pd.DataFrame(rows)


def talkers_frame(stats: dict) -> "pd.DataFrame":
    return pd.DataFrame(stats.get("top_talkers", []))


def evidence_frame(stats: dict) -> "pd.DataFrame":
    """The two things the brief asks Phase 2 to identify explicitly: a TCP
    three-way handshake and a DNS lookup. Kept on their own sheet so they are
    easy to point at during the presentation."""
    rows = []
    for handshake in stats.get("tcp_handshakes", [])[:15]:
        rows.append({
            "type": "TCP three-way handshake",
            "detail": f"{handshake.get('client')}:{handshake.get('client_port')} -> "
                      f"{handshake.get('server')}:{handshake.get('server_port')} "
                      f"({handshake.get('service', '')})",
            "frames": f"SYN #{handshake.get('syn', {}).get('frame')} -> "
                      f"SYN/ACK #{handshake.get('syn_ack', {}).get('frame')} -> "
                      f"ACK #{handshake.get('ack', {}).get('frame')}",
            "timestamp": handshake.get("syn", {}).get("timestamp"),
        })
    for lookup in stats.get("dns_lookups", [])[:15]:
        rows.append({
            "type": "DNS lookup",
            "detail": f"{lookup.get('client')} asked {lookup.get('server')} "
                      f"for {lookup.get('query')}",
            "frames": f"#{lookup.get('frame')}",
            "timestamp": lookup.get("timestamp"),
        })
    return pd.DataFrame(rows)


def mac_frame(mac_log: dict | None) -> "pd.DataFrame":
    if not mac_log:
        return pd.DataFrame()
    rows = [{
        "stage": entry.get("stage"),
        "timestamp": entry.get("timestamp"),
        "host": entry.get("host"),
        "interface": entry.get("interface"),
        "mac": entry.get("mac"),
        "ipv4": entry.get("ipv4"),
        "note": entry.get("note"),
    } for entry in mac_log.get("entries", [])]
    return pd.DataFrame(rows)


def summary_frame(summary: dict, data: dict) -> "pd.DataFrame":
    counts = summary.get("findings_by_severity", {})
    stats = data.get("stats", {})
    rows = [
        ("Report generated", summary.get("generated_at")),
        ("Scope", summary.get("scope")),
        ("", ""),
        ("Hosts assessed (Phase 1)", summary.get("hosts_assessed", 0)),
        ("Open ports found (Phase 1)", summary.get("open_ports_found", 0)),
        ("Packets analysed (Phase 2)", summary.get("packets_analysed", 0)),
        ("Capture duration (s)", stats.get("duration_seconds")),
        ("Distinct protocols seen (Phase 2)", summary.get("protocols_seen", 0)),
        ("TCP handshakes identified", stats.get("tcp_handshakes_total", 0)),
        ("DNS lookups identified", stats.get("dns_lookups_total", 0)),
        ("MAC spoofing demonstrated (Phase 3)",
         "yes" if summary.get("mac_spoofing_demonstrated") else "not recorded"),
        ("", ""),
        ("Total findings", summary.get("findings_total", 0)),
    ]
    rows += [(f"  {severity} findings", counts.get(severity, 0))
             for severity in SEVERITY_ORDER]
    rows += [("", ""), ("Top risks", "")]
    rows += [(f"  {i}.", title)
             for i, title in enumerate(summary.get("top_risks", []), start=1)]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


# --------------------------------------------------------------------------
# Charts
# --------------------------------------------------------------------------


def style_axes(axes, title: str, xlabel: str = "", ylabel: str = "") -> None:
    axes.set_title(title, fontsize=13, fontweight="bold", color=ACCENT, pad=14)
    axes.set_xlabel(xlabel, fontsize=10)
    axes.set_ylabel(ylabel, fontsize=10)
    axes.spines["top"].set_visible(False)
    axes.spines["right"].set_visible(False)
    axes.grid(axis="y", alpha=0.25, linestyle="--")
    axes.set_axisbelow(True)


def save(figure, path: Path) -> Path:
    ensure_dir(path.parent)
    figure.tight_layout()
    figure.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(figure)
    ok(f"chart: {rel(path)}")
    return path


def chart_severity(findings: "pd.DataFrame", directory: Path) -> Path | None:
    if findings.empty:
        return None
    counts = findings["severity"].value_counts().reindex(SEVERITY_ORDER, fill_value=0)
    figure, axes = plt.subplots(figsize=(7, 4.2))
    bars = axes.bar(counts.index, counts.values,
                    color=[SEVERITY_COLOURS[s] for s in counts.index])
    for bar, value in zip(bars, counts.values):
        if value:
            axes.text(bar.get_x() + bar.get_width() / 2, value + 0.15, str(value),
                      ha="center", fontsize=10, fontweight="bold")
    style_axes(axes, "Findings by severity", ylabel="findings")
    axes.set_ylim(0, max(counts.values.max() * 1.18, 1))
    return save(figure, directory / "findings_by_severity.png")


def chart_open_ports(hosts: "pd.DataFrame", directory: Path) -> Path | None:
    if hosts.empty or hosts["open_ports"].sum() == 0:
        return None
    labels = [f"{row.ip}\n{row.team_host}" if row.team_host else str(row.ip)
              for row in hosts.itertuples()]
    figure, axes = plt.subplots(figsize=(7, 4.2))
    bars = axes.bar(labels, hosts["open_ports"], color=ACCENT)
    for bar, value in zip(bars, hosts["open_ports"]):
        axes.text(bar.get_x() + bar.get_width() / 2, value + 0.08, str(value),
                  ha="center", fontsize=10, fontweight="bold")
    style_axes(axes, "Open TCP ports per host", ylabel="open ports")
    axes.tick_params(axis="x", labelsize=8)
    return save(figure, directory / "open_ports_by_host.png")


def chart_services(ports: "pd.DataFrame", directory: Path) -> Path | None:
    if ports.empty:
        return None
    open_ports = ports[ports["state"] == "open"]
    if open_ports.empty:
        return None
    counts = (open_ports.assign(
        label=open_ports["service"].fillna("unknown") + " (" +
              open_ports["port"].astype(str) + ")")
        ["label"].value_counts().head(12).sort_values())
    figure, axes = plt.subplots(figsize=(7, max(3.2, 0.42 * len(counts) + 1.4)))
    axes.barh(counts.index, counts.values, color=ACCENT)
    style_axes(axes, "Exposed services across the network", xlabel="hosts exposing it")
    axes.grid(axis="x", alpha=0.25, linestyle="--")
    axes.grid(axis="y", visible=False)
    return save(figure, directory / "top_services.png")


def chart_protocols(protocols: "pd.DataFrame", directory: Path) -> Path | None:
    if protocols.empty:
        return None
    top = protocols.sort_values("packets", ascending=False).head(8)
    others = protocols["packets"].sum() - top["packets"].sum()
    labels = list(top["protocol"])
    values = list(top["packets"])
    if others > 0:
        labels.append("other")
        values.append(int(others))

    figure, axes = plt.subplots(figsize=(6.4, 5))
    colours = plt.cm.tab20.colors[:len(values)]
    wedges, _, autotexts = axes.pie(
        values, labels=labels, autopct="%1.1f%%", startangle=110,
        colors=colours, pctdistance=0.78,
        wedgeprops={"width": 0.55, "edgecolor": "white", "linewidth": 1.5},
        textprops={"fontsize": 9},
    )
    for text in autotexts:
        text.set_fontsize(8)
    axes.set_title("Captured traffic by protocol", fontsize=13,
                   fontweight="bold", color=ACCENT, pad=16)
    return save(figure, directory / "protocol_distribution.png")


def chart_talkers(talkers: "pd.DataFrame", directory: Path) -> Path | None:
    if talkers.empty or "packets_sent" not in talkers:
        return None
    top = talkers.head(10).iloc[::-1]
    labels = [f"{row.ip}  {row.label}" if getattr(row, "label", None) else str(row.ip)
              for row in top.itertuples()]
    figure, axes = plt.subplots(figsize=(7.5, max(3.2, 0.42 * len(top) + 1.4)))
    axes.barh(labels, top["packets_sent"], color=ACCENT)
    style_axes(axes, "Top talkers (packets sent)", xlabel="packets")
    axes.grid(axis="x", alpha=0.25, linestyle="--")
    axes.grid(axis="y", visible=False)
    axes.tick_params(axis="y", labelsize=8)
    return save(figure, directory / "top_talkers.png")


def build_charts(frames: dict, directory: Path) -> list[Path]:
    step("Building charts")
    ensure_dir(directory)
    charts = [
        chart_severity(frames["findings"], directory),
        chart_open_ports(frames["hosts"], directory),
        chart_services(frames["ports"], directory),
        chart_protocols(frames["protocols"], directory),
        chart_talkers(frames["talkers"], directory),
    ]
    return [c for c in charts if c]


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------


HEADER_FILL = PatternFill("solid", fgColor="2F4858")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def autosize(worksheet, frame: "pd.DataFrame", max_width: int = 60) -> None:
    for index, column in enumerate(frame.columns, start=1):
        values = frame[column].astype(str)
        longest = max([len(str(column))] + [len(v) for v in values[:200]], default=10)
        worksheet.column_dimensions[get_column_letter(index)].width = \
            min(max(longest + 2, 10), max_width)


def style_header(worksheet, columns: int) -> None:
    for index in range(1, columns + 1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    worksheet.row_dimensions[1].height = 20
    worksheet.freeze_panes = "A2"


def colour_severity(worksheet, frame: "pd.DataFrame") -> None:
    if "severity" not in frame.columns:
        return
    column = list(frame.columns).index("severity") + 1
    for row_index, severity in enumerate(frame["severity"], start=2):
        fill = SEVERITY_FILLS.get(str(severity))
        if fill:
            for col in range(1, len(frame.columns) + 1):
                worksheet.cell(row=row_index, column=col).fill = \
                    PatternFill("solid", fgColor=fill)
        cell = worksheet.cell(row=row_index, column=column)
        cell.font = Font(bold=True)


def wrap_columns(worksheet, frame: "pd.DataFrame", columns: list[str],
                 width: int = 55) -> None:
    for name in columns:
        if name not in frame.columns:
            continue
        index = list(frame.columns).index(name) + 1
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = width
        for row in range(2, len(frame) + 2):
            worksheet.cell(row=row, column=index).alignment = \
                Alignment(wrap_text=True, vertical="top")


def _write_activity_sheets(writer) -> None:
    """Add the employee network-activity monitor sheets (the monitoring lens)
    to the workbook when activity_monitor.py has produced activity.json.
    Restructures the Phase 4 deliverable to carry both the security audit and
    the monitoring narrative in one book."""
    path = config.PHASE4_OUTPUTS / "activity.json"
    if not path.exists():
        return
    try:
        report = read_json(path)
    except (ValueError, OSError):
        return

    # 1. Devices (one row per monitored employee/device)
    dev_rows = []
    for e in report.get("employees", []):
        act = e.get("activity", {})
        dev_rows.append({
            "device": e.get("label"),
            "identified": "yes" if e.get("identified") else "no",
            "ip": e.get("ip"), "mac": e.get("mac"),
            "traffic_bytes": e.get("bytes"),
            "active_minutes": e.get("active_minutes"),
            "unique_domains": e.get("unique_domains"),
            "activity_score*": act.get("headline_score"),
            "work_site_ratio_%": act.get("work_domain_ratio"),
            "presence_%": act.get("presence_pct"),
            "top_categories": ", ".join(f"{k}:{v}" for k, v in
                                        list(e.get("category_breakdown", {}).items())[:4]),
            "flags": " | ".join(e.get("flags", [])),
        })
    devices = pd.DataFrame(dev_rows) if dev_rows else pd.DataFrame({"(no devices)": []})

    # 2. Domains (org-wide)
    dom_rows = [{"domain": d.get("domain"), "hits": d.get("hits"),
                 "category": d.get("category")}
                for d in report.get("organisation", {}).get("top_domains", [])]
    domains = pd.DataFrame(dom_rows) if dom_rows else pd.DataFrame({"(no domains)": []})

    # 3. Blind spots + scope/privacy notes
    bs = report.get("blind_spots", {})
    blind_rows = [{"blind spot / note": n} for n in bs.get("notes", [])]
    blind_rows += [
        {"blind spot / note": ""},
        {"blind spot / note": "SCOPE: " + report.get("scope_caveat", "")},
        {"blind spot / note": "PRIVACY: " + report.get("privacy_notice", "")},
    ]
    blind = pd.DataFrame(blind_rows)

    for name, frame in (("Activity Devices", devices),
                        ("Activity Domains", domains),
                        ("Blind Spots", blind)):
        frame.to_excel(writer, sheet_name=name, index=False)
        ws = writer.sheets[name]
        style_header(ws, len(frame.columns))
        autosize(ws, frame)
        if name == "Blind Spots":
            ws.column_dimensions["A"].width = 105
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ok("added monitoring sheets (Activity Devices / Domains / Blind Spots)")


def build_workbook(frames: dict, data: dict, charts: list[Path],
                   path: Path) -> Path:
    step("Writing the workbook")
    ensure_dir(path.parent)

    sheets: list[tuple[str, "pd.DataFrame"]] = [
        ("Summary", frames["summary"]),
        ("Findings", frames["findings"]),
        ("Hosts", frames["hosts"]),
        ("Open Ports", frames["ports"]),
        ("Protocol Stats", frames["protocols"]),
        ("Top Talkers", frames["talkers"]),
        ("Evidence", frames["evidence"]),
    ]
    if not frames["mac"].empty:
        sheets.append(("MAC Log", frames["mac"]))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets:
            output = frame if not frame.empty else pd.DataFrame({"(no data)": []})
            output.to_excel(writer, sheet_name=name, index=False)

        # Firewall rules go in as text, one line per row.
        if config.FIREWALL_RULES_TXT.exists():
            lines = config.FIREWALL_RULES_TXT.read_text(encoding="utf-8").splitlines()
            pd.DataFrame({"Recommended firewall configuration": lines}).to_excel(
                writer, sheet_name="Firewall Rules", index=False)

        workbook = writer.book
        for name, frame in sheets:
            worksheet = writer.sheets[name]
            if frame.empty:
                continue
            style_header(worksheet, len(frame.columns))
            autosize(worksheet, frame)
            if name == "Findings":
                colour_severity(worksheet, frame)
                wrap_columns(worksheet, frame,
                             ["evidence", "risk", "recommendation"], width=52)
                wrap_columns(worksheet, frame, ["title"], width=34)
            if name == "Summary":
                worksheet.column_dimensions["A"].width = 38
                worksheet.column_dimensions["B"].width = 72
                for row in range(2, len(frame) + 2):
                    worksheet.cell(row=row, column=2).alignment = \
                        Alignment(wrap_text=True, vertical="top")

        if "Firewall Rules" in writer.sheets:
            worksheet = writer.sheets["Firewall Rules"]
            style_header(worksheet, 1)
            worksheet.column_dimensions["A"].width = 110
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=1).font = Font(name="Consolas", size=9)

        # Monitoring lens: employee network-activity sheets, if produced.
        _write_activity_sheets(writer)

        # Charts on their own sheet so the deck can screenshot straight from it.
        if charts:
            sheet = workbook.create_sheet("Charts")
            sheet["A1"] = "Charts (also saved as PNGs in outputs/charts/)"
            sheet["A1"].font = Font(bold=True, size=12, color="2F4858")
            row = 3
            for chart in charts:
                image = XLImage(str(chart))
                # ~0.75pt per px; keep the sheet readable rather than exact.
                scale = min(1.0, 620 / image.width)
                image.width = int(image.width * scale)
                image.height = int(image.height * scale)
                sheet.add_image(image, f"A{row}")
                row += int(image.height / 19) + 3

    ok(f"wrote {rel(path)}")
    return path


# --------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="Phase 4 - charts and Excel report")
    parser.add_argument("--no-charts", action="store_true",
                        help="skip chart generation (workbook only)")
    parser.add_argument("--output", type=Path, default=config.REPORT_XLSX)
    args = parser.parse_args()

    banner("PHASE 4 - REPORT")
    step("Loading analysis output")
    data = load_everything()
    info(f"{len(data['findings'])} finding(s), {len(data['hosts'])} host(s), "
         f"{data['stats'].get('total_packets', 0)} packet(s)")

    frames = {
        "summary": summary_frame(data["summary"], data),
        "findings": findings_frame(data["findings"]),
        "hosts": hosts_frame(data["hosts"]),
        "ports": ports_frame(data["hosts"]),
        "protocols": protocol_frame(data["stats"]),
        "talkers": talkers_frame(data["stats"]),
        "evidence": evidence_frame(data["stats"]),
        "mac": mac_frame(data["mac_log"]),
    }

    charts: list[Path] = []
    if not args.no_charts:
        charts = build_charts(frames, config.CHARTS_DIR)
        if not charts:
            warn("no charts produced - the inputs had nothing to plot")

    build_workbook(frames, data, charts, Path(args.output))

    step("Deliverables")
    info(f"report   : {rel(Path(args.output))}")
    info(f"charts   : {rel(config.CHARTS_DIR)}")
    if config.FIREWALL_RULES_TXT.exists():
        info(f"firewall : {rel(config.FIREWALL_RULES_TXT)}")
    if data["is_sample"]:
        warn("built from sample data - re-run analyze_security.py without "
             "--sample once Phases 1 and 2 deliver.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print()
        warn("interrupted")
        raise SystemExit(130)
